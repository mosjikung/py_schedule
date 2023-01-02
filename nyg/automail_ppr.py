#-*-coding: utf-8 -*-
import tkinter as tk
from tkinter import messagebox
import cx_Oracle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import datetime
import threading
import smtplib
import ssl
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"


def gen_cursor_to_file(ws, cursor):
  for c, col in enumerate(cursor.iteritems()):
    ws.cell(row=1, column=c+1, value=col[0])

  r = 2
  for index, row in cursor.iterrows():
    for c, col in enumerate(cursor.iteritems()):
      ws.cell(row=r, column=c+1,
              value=str(row[col[0]]).encode("ascii", errors="ignore"))
    r = r + 1



class CLS_AutoMail_PPR(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      AutoMail_PPR()
      AutoMail_PPR_SUB()
      SendMailPPR()




def AutoMail_PPR():
  FileNameXLSX = 'QN-Data-NoColor.xlsx'
  time_start = datetime.now()
  my_dsn = cx_Oracle.makedsn("172.16.6.83", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()

  cursor.execute(""" 
  SELECT M.* FROM (
SELECT BU,
GMT_TYPE,
STYLE_REF,
SO_NO,
SO_YEAR,
SO_NO_DOC,
TO_CHAR(CREATE_DATE,'RRRR-MM-DD') AS SO_DATE,
ORDER_TYPE,
ORDER_TYPE_DESC,
MIN(TO_CHAR(SHIPMENT_DATE,'RRRR-MM-DD')) SHIPMENT_DATE,
CUST_GROUP,
CUST_NAME,
BRAND_NAME,
MIN(EDD_W) EDD_W,
MIN(RDD_W) RDD_W,
--MIN(EDD_W_NEW) AS WK_EDD_MINUS_4,
--MIN(RDD_W_NEW) AS WK_RDD_MINUS_4,
MIN(MRD_W) AS MRD_W,
SUM(ORDER_QTY) ORDER_QTY,
TO_CHAR(SO_RELEASE_DATE,'RRRR-MM-DD') SO_RELEASE_DATE,
TO_CHAR(SAMPLE_RELEASE_DATE,'RRRR-MM-DD') SAMPLE_RELEASE_DATE,
TO_CHAR(FABRIC_COMPLETE_DATE,'RRRR-MM-DD') FABRIC_COMPLETE_DATE,
TO_CHAR(SEW_ACC_COMPLETE_DATE,'RRRR-MM-DD') SEW_ACC_COMPLETE_DATE,
TO_CHAR(PACK_ACC_COMPLETE_DATE,'RRRR-MM-DD') PACK_ACC_COMPLETE_DATE,
TO_CHAR(FC_RELEASE_DATE,'RRRR-MM-DD') FC_RELEASE_DATE,
TO_CHAR(PATTERN_RELEASE_DATE,'RRRR-MM-DD') PATTERN_RELEASE_DATE,
SCM_ALLOCATE,
TO_CHAR(FIRST_CUT_DATE,'RRRR-MM-DD') FIRST_CUT_DATE,
--SO_READY,
ACT_CUT,
LOADED_QTY,
SUM(ORDER_QTY) - LOADED_QTY AS PENDING
FROM CONTROL_REP_PPR_WK M
WHERE 1 = 1

GROUP BY BU,
GMT_TYPE,
STYLE_REF,
SO_NO,
SO_YEAR,
SO_NO_DOC,
TO_CHAR(CREATE_DATE,'RRRR-MM-DD'),
ORDER_TYPE,
ORDER_TYPE_DESC,
CUST_GROUP,
CUST_NAME,
BRAND_NAME,
TO_CHAR(SO_RELEASE_DATE,'RRRR-MM-DD') ,
TO_CHAR(SAMPLE_RELEASE_DATE,'RRRR-MM-DD') ,
TO_CHAR(FABRIC_COMPLETE_DATE,'RRRR-MM-DD') ,
TO_CHAR(SEW_ACC_COMPLETE_DATE,'RRRR-MM-DD') ,
TO_CHAR(PACK_ACC_COMPLETE_DATE,'RRRR-MM-DD') ,
TO_CHAR(FC_RELEASE_DATE,'RRRR-MM-DD') ,
TO_CHAR(PATTERN_RELEASE_DATE,'RRRR-MM-DD') ,
SCM_ALLOCATE,
TO_CHAR(FIRST_CUT_DATE,'RRRR-MM-DD') ,
--SO_READY,
ACT_CUT,
LOADED_QTY ) M
WHERE MRD_W >= '202025'
ORDER BY MRD_W, SO_NO_DOC 
   """)


  wb = openpyxl.Workbook()
  ws = wb['Sheet']

  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])

  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
    r = r + 1
  fullpath = r'C:\GITProject\pyschedule\AutoMail_PPR.xlsx'
  wb.save(fullpath)

  conn.close()
  
  print('Complete AutoMail_PPR')


def AutoMail_PPR_SUB():
  FileNameXLSX = 'QN-Data-NoColor.xlsx'
  time_start = datetime.now()
  my_dsn = cx_Oracle.makedsn("172.16.6.83", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  # cursor = conn.cursor()

  sql = """ 
   SELECT BU,
GMT_TYPE,
STYLE_REF,
SO_NO,
SO_YEAR,
SUB_NO,
SO_NO_DOC,
 TO_CHAR(CREATE_DATE,'DD-MON-YY')  AS SO_DATE,
ORDER_TYPE,
ORDER_TYPE_DESC,
TO_CHAR(SHIPMENT_DATE,'DD-MON-YY')  SHIPMENT_DATE,
CUST_GROUP,
CUST_NAME,
BRAND_NAME,
EDD_W,
RDD_W,
MRD_W ,
ORDER_ID,
COLOR_FC,
ORDER_QTY,
TO_CHAR(SO_RELEASE_DATE,'DD-MON-YY') SO_RELEASE_DATE,
TO_CHAR(SAMPLE_RELEASE_DATE,'DD-MON-YY') SAMPLE_RELEASE_DATE,
TO_CHAR(FABRIC_COMPLETE_DATE,'DD-MON-YY') FABRIC_COMPLETE_DATE,
TO_CHAR(SEW_ACC_COMPLETE_DATE,'DD-MON-YY') SEW_ACC_COMPLETE_DATE,
TO_CHAR(PACK_ACC_COMPLETE_DATE,'DD-MON-YY') PACK_ACC_COMPLETE_DATE,
TO_CHAR(FC_RELEASE_DATE,'DD-MON-YY') FC_RELEASE_DATE,
TO_CHAR(PATTERN_RELEASE_DATE,'DD-MON-YY') PATTERN_RELEASE_DATE,
SCM_ALLOCATE,
TO_CHAR(FIRST_CUT_DATE,'DD-MON-YY') FIRST_CUT_DATE,
ACT_CUT
,LOADED_QTY
	FROM CONTROL_REP_PPR_WK M
	WHERE MRD_W >= '202025' 
ORDER BY MRD_W, SO_NO_DOC, SUB_NO 
   """

  df = pd.read_sql_query(sql, conn)
  df.fillna("", inplace=True)


  df = df.applymap(lambda x: x.encode('unicode_escape').
                   decode('utf-8') if isinstance(x, str) else x)

                   

  # print(df)

  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  home = os.path.expanduser('~')

  gen_cursor_to_file(ws, df)
  path = r"C:\GITProject\pyschedule\AutoMail_PPR_SUB.xlsx"
  wb.save(path)

  conn.close()
  print('Complete AutoMail_PPR_SUB')


def SendMailPPR():
    msg = MIMEMultipart()
    # message = "Production send mail to \r\r\n" + mailto + "\r\r\n Thank you"
    message = "\r\r\n Thank you"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

    recipients = "kaweewat.k@nanyangtextile.com,penpayom.m@nanyangtextile.com"
    # bcc = 'piyawat.k@nanyangtextile.com,kaweewat.k@nanyangtextile.com,somchai.c@nanyangtextile.com'
    # recipients = 'kaweewat.k@nanyangtextile.com'
    # bcc = 'kaweewat.k@gmail.com'

    msg['From'] = "admin_monitoring.k@nanyangtextile.com"
    msg['To'] = recipients
    
    msg['Subject'] = 'Auto Mail PPR'

    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    # msg.attach(MIMEText(message, 'plain'))

    filename = r'AutoMail_PPR.xlsx'
    attachment = open(r'C:\GITProject\pyschedule\AutoMail_PPR.xlsx', 'rb')
    xlsx = MIMEBase(
        'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    xlsx.set_payload(attachment.read())

    encoders.encode_base64(xlsx)
    xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(xlsx)

    filename2 = r'AutoMail_PPR_SUB.xlsx'
    attachment2 = open(r'C:\GITProject\pyschedule\AutoMail_PPR_SUB.xlsx', 'rb')
    xlsx2 = MIMEBase(
        'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    xlsx2.set_payload(attachment2.read())

    encoders.encode_base64(xlsx2)
    xlsx2.add_header('Content-Disposition', 'attachment', filename=filename2)
    msg.attach(xlsx2)

    rcpt = recipients.split(',')

    server = smtplib.SMTP('smtp.nanyangtextile.com: 25')

    server.sendmail(msg['From'], rcpt, msg.as_string())
    server.quit()
    print("Send Mail PPR")


############################################
threads = []

thread1 = CLS_AutoMail_PPR()
thread1.start()
threads.append(thread1)

for t in threads:
    t.join()
print("COMPLETE")
