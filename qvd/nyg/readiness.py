import cx_Oracle
import csv
import os
from pathlib import Path
import requests
from datetime import datetime
import threading
import time
import openpyxl
import mysql.connector
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"]=oracle_client
os.environ["PATH"]=oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"]="AMERICAN_AMERICA.TH8TISASCII"


def gen_cursor_to_file(PathFileNameXLSX, cursor):
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  home = os.path.expanduser('~')
  
  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])
  
  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
    r = r + 1
  wb.save(PathFileNameXLSX)


###########################################
class CLS_CONTROL_WIP_READINESS_ALL1(threading.Thread):
      def __init__(self):
        threading.Thread.__init__(self)
      def run(self):
        CONTROL_WIP_READINESS_ALL1()
        
def CONTROL_WIP_READINESS_ALL1():
  my_dsn = cx_Oracle.makedsn("172.16.6.83",port=1521,sid="NYTG")
  conn = cx_Oracle.connect(user="NYGM", password="NYGM", dsn=my_dsn, encoding = "UTF-8", nencoding = "UTF-8")
  cursor = conn.cursor()
  cursor.execute("""select BU, GMT_TYPE, STYLE_REF, SO_NO, SO_YEAR, SUB_NO, SO_NO_DOC, TO_CHAR(SHIPMENT_DATE,'YYYY-MM-DD') SHIPMENT_DATE, CUST_GROUP, CUST_NAME
, BRAND_NAME, EDD_W, ORDER_ID, COLOR_FC, nvl(ORDER_QTY,0) ORDER_QTY, SO_RELEASE, TO_CHAR(SO_RELEASE_DATE,'YYYY-MM-DD') SO_RELEASE_DATE, SAMPLE_RELEASE
, TO_CHAR(SAMPLE_RELEASE_DATE,'YYYY-MM-DD') SAMPLE_RELEASE_DATE, PATTERN_RELEASE, TO_CHAR(PATTERN_RELEASE_DATE,'YYYY-MM-DD') PATTERN_RELEASE_DATE, FC_RELEASE
, TO_CHAR(FC_RELEASE_DATE,'YYYY-MM-DD') FC_RELEASE_DATE
, SEW_ACC_COMPLETE, TO_CHAR(SEW_ACC_COMPLETE_DATE,'YYYY-MM-DD') SEW_ACC_COMPLETE_DATE, PACK_ACC_COMPLETE, TO_CHAR(PACK_ACC_COMPLETE_DATE,'YYYY-MM-DD') PACK_ACC_COMPLETE_DATE
, FABRIC_COMPLETE, TO_CHAR(FABRIC_COMPLETE_DATE,'YYYY-MM-DD')  FABRIC_COMPLETE_DATE, SCM_ALLOCATE, ORDER_TYPE, ORDER_TYPE_DESC, RDD_W, FIRST_CUT_DATE
, CASE WHEN m.so_release = 'Y' and m.sample_release = 'Y' and m.fabric_complete = 'Y' 
and m.fc_release = 'Y' and m.sew_acc_complete = 'Y' and m.pack_acc_complete = 'Y'
and m.fc_release = 'Y' and m.pattern_release = 'Y' THEN 'Y' ELSE 'N' END SO_READY
from OE_SUB_READINESS_STATUS_V m
where first_cut_date is null 
and nvl(order_qty,0) > 0 
ORDER BY 35 DESC, SHIPMENT_DATE

""")

# and m.so_release = 'Y'
# and m.sample_release = 'Y' 
# and m.fabric_complete = 'Y'
# and m.fc_release = 'Y'
# and m.sew_acc_complete = 'Y'
# and m.pack_acc_complete = 'Y'
# and m.fc_release = 'Y'
# and m.pattern_release = 'Y'

  print("Gen File CLS_CONTROL_WIP_READINESS_ALL1")
  path = r"C:\IT_ONLY\NYGCONTROLROOM\CONTROL_WIP_READINESS_ALL1.xlsx"
  gen_cursor_to_file(path, cursor)

  conn.close()
  print('Complete CLS_CONTROL_WIP_READINESS_ALL1')


###########################################
class CLS_CONTROL_WIP_READINESS_ALL2(threading.Thread):
      def __init__(self):
        threading.Thread.__init__(self)
      def run(self):
        CONTROL_WIP_READINESS_ALL2()
        
def CONTROL_WIP_READINESS_ALL2():
  my_dsn = cx_Oracle.makedsn("172.16.6.83",port=1521,sid="NYTG")
  conn = cx_Oracle.connect(user="NYGM", password="NYGM", dsn=my_dsn, encoding = "UTF-8", nencoding = "UTF-8")
  cursor = conn.cursor()
  cursor.execute("""select BU, GMT_TYPE, STYLE_REF, SO_NO, SO_YEAR, SUB_NO, SO_NO_DOC, TO_CHAR(SHIPMENT_DATE,'YYYY-MM-DD') SHIPMENT_DATE, CUST_GROUP, CUST_NAME
, BRAND_NAME, EDD_W, ORDER_ID, COLOR_FC, nvl(ORDER_QTY,0) ORDER_QTY, SO_RELEASE, TO_CHAR(SO_RELEASE_DATE,'YYYY-MM-DD') SO_RELEASE_DATE, SAMPLE_RELEASE
, TO_CHAR(SAMPLE_RELEASE_DATE,'YYYY-MM-DD') SAMPLE_RELEASE_DATE, PATTERN_RELEASE, TO_CHAR(PATTERN_RELEASE_DATE,'YYYY-MM-DD') PATTERN_RELEASE_DATE, FC_RELEASE
, TO_CHAR(FC_RELEASE_DATE,'YYYY-MM-DD') FC_RELEASE_DATE
, SEW_ACC_COMPLETE, TO_CHAR(SEW_ACC_COMPLETE_DATE,'YYYY-MM-DD') SEW_ACC_COMPLETE_DATE, PACK_ACC_COMPLETE, TO_CHAR(PACK_ACC_COMPLETE_DATE,'YYYY-MM-DD') PACK_ACC_COMPLETE_DATE
, FABRIC_COMPLETE, TO_CHAR(FABRIC_COMPLETE_DATE,'YYYY-MM-DD')  FABRIC_COMPLETE_DATE, SCM_ALLOCATE, ORDER_TYPE, ORDER_TYPE_DESC, RDD_W, FIRST_CUT_DATE
, CASE WHEN m.so_release = 'Y' and m.sample_release = 'Y' and m.fabric_complete = 'Y' 
and m.fc_release = 'Y' and m.sew_acc_complete = 'Y' and m.pack_acc_complete = 'Y'
and m.fc_release = 'Y' and m.pattern_release = 'Y' THEN 'Y' ELSE 'N' END SO_READY
from OE_SUB_READINESS_STATUS_V m
where first_cut_date is null 
and nvl(order_qty,0) > 0
ORDER BY 35 DESC, SHIPMENT_DATE

""")


# and m.so_release = 'Y'
# and m.sample_release = 'Y' 
# and m.fabric_complete = 'Y'
# and m.fc_release = 'Y'
# and m.sew_acc_complete = 'Y'
# and m.pack_acc_complete = 'Y'
# and m.fc_release = 'Y'
# and m.pattern_release = 'Y'

  print("Gen File CLS_CONTROL_WIP_READINESS_ALL2")
  path = r"C:\IT_ONLY\NYGCONTROLROOM\CONTROL_WIP_READINESS_ALL2.xlsx"
  gen_cursor_to_file(path, cursor)

  conn.close()
  print('Complete CLS_CONTROL_WIP_READINESS_ALL2')


threads = []

thread1 = CLS_CONTROL_WIP_READINESS_ALL1();thread1.start();threads.append(thread1)
thread2 = CLS_CONTROL_WIP_READINESS_ALL2();thread2.start();threads.append(thread2)
# thread3 = CLS_peopleVN();thread3.start();threads.append(thread3)



for t in threads:
    t.join()
print ("COMPLETE All")
