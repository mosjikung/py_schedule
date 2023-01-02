import os
import io
import cx_Oracle
import numpy as np
import pandas as pd
import json
# from flask import jsonify
from pandas.io.json import json_normalize
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

oracle_client = r"C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"

my_dsn = cx_Oracle.makedsn("172.16.6.83", port=1521, sid="NYTG")

def gen_cursor_to_file(ws, cursor):
  green = '33FF99'
  red = 'FF9999'
  for c, col in enumerate(cursor.iteritems()):
    ws.cell(row=1, column=c+1, value=col[0])

  r = 2
  for index, row in cursor.iterrows():
    for c, col in enumerate(cursor.iteritems()):
      ws.cell(row=r, column=c+1, value=row[col[0]])
      if(row['OTP']=='HIT'):
        ws.cell(row=r, column=c+1).fill = PatternFill(start_color=green,
                                                 end_color=green,
                                                   fill_type='solid')
      else:
        ws.cell(row=r, column=c+1).fill = PatternFill(start_color=red,
                                                 end_color=red,
                                                   fill_type='solid')
    r = r + 1

def otp_doc_grp(grp, targetwk, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board1_otif m
            where grp = '{grp}'
            and m.target_by_rdd = '{targetwk}'
            """.format(targetwk = targetwk, grp = grp)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Document of {grp} Week = {targetwk}".format(grp=grp, targetwk=targetwk))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_doc_otp_{grp}_ALL_{targetwk}_{uuid}.xlsx".format(grp=grp, targetwk=targetwk, uuid=uuid)
  wb.save(path)
  
  
  
def otp_sam_grp(grp, targetwk, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board2_otif m
            where grp = '{grp}'
            and m.target_by_rdd = '{targetwk}'
            """.format(targetwk = targetwk, grp = grp)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Sample of {grp} Week = {targetwk}".format(grp=grp, targetwk=targetwk))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_sample_otp_{grp}_ALL_{targetwk}_{uuid}.xlsx".format(grp=grp, targetwk=targetwk, uuid=uuid)
  wb.save(path)


def otp_doc_team(team, targetwk, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board1_otif m
            where grp || '-' || loc = '{team}'
            and m.target_by_rdd = '{targetwk}'
            """.format(targetwk = targetwk, team = team)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Document of {team} Week = {targetwk}".format(team=team, targetwk=targetwk))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_doc_otp_{team}_ALL_{targetwk}_{uuid}.xlsx".format(team=team, targetwk=targetwk, uuid=uuid)
  wb.save(path)


def otp_sample_team(team, targetwk, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board2_otif m
            where grp || '-' || loc = '{team}'
            and m.target_by_rdd = '{targetwk}'
            """.format(targetwk = targetwk, team = team)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Sample of {team} Week = {targetwk}".format(team=team, targetwk=targetwk))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_sample_otp_{team}_ALL_{targetwk}_{uuid}.xlsx".format(team=team, targetwk=targetwk, uuid=uuid)
  wb.save(path)

# otp_sample_team('NIKE-TH NAY1-1', '202050', 'xxx')

# otp_doc_grp('ALL', '202101', 'uuid')

def fon_doc(grp, ws, we, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board1_otif m
            where 1 = 1
            and m.target_by_rdd >= '{ws}'
            and m.target_by_rdd <= '{we}'
            """.format(ws = ws, we=we, grp = grp)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Document of {grp} ".format(grp=grp))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_doc_otp_{grp}_ALL.xlsx".format(grp=grp)
  wb.save(path)
  
  
def fon_sample(grp, ws, we, uuid):
    
  conn = cx_Oracle.connect(user="nygm", password="nygm",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  
  sql = """ select RANK() OVER (ORDER BY grp, loc, so_no_doc, SHIPMENT_DATE) rnk, m.*
            ,case when nvl(actual_wk,999999) > target_by_rdd then 'FAIL' else 'HIT' end OTP
            from ppr_board2_otif m
            where 1 = 1
            and m.target_by_rdd >= '{ws}'
            and m.target_by_rdd <= '{we}'
            """.format(ws = ws, we=we, grp = grp)
            
  dfso = pd.read_sql_query(sql, conn)
  # print(dfso)
  
  wb = openpyxl.Workbook()
  ws = wb['Sheet']
  
  ws.cell(row=1, column=1, value="OTP Sample of {grp} ".format(grp=grp))
  
  
  gen_cursor_to_file(ws, dfso)
  
  conn.close()
  path = r"C:\\flask\\static\\so_sample_otp_{grp}_ALL.xlsx".format(grp=grp,uuid=uuid)
  wb.save(path)
  


fon_doc('ALL','202048','202116','')
fon_sample('ALL','202048','202116','')