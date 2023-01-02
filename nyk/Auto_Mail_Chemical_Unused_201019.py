#-*-coding: utf-8 -*-
import tkinter as tk
from tkinter import messagebox
import cx_Oracle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import datetime
import threading
import smtplib
import ssl
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"


class CLS_ExPortQnNoColor(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      ExPortQnNoColor()

def ExPortQnNoColor():
  P_ActiveQ1 = 0
  P_ActiveQ2 = 0
  FileNameXLSX = 'Chemical_Unused.xlsx'
  time_start = datetime.now()

  my_dsn = cx_Oracle.makedsn("172.16.6.76", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="demo", password="demo",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" Select Formula_No,Chemical_Code,Chemical_Desc,Concentrated,ITEMS_REPLACE,ITEMS_REPLACE_DESC,LAST_UPD_DATE From V_ICemical_formula_unused """)

  wb = openpyxl.Workbook()
  ws = wb['Sheet']

  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])
    P_ActiveQ1 = 1

  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
    r = r + 1
    P_ActiveQ2 = P_ActiveQ2 + 1
  #fullpath = r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx'
  fullpath = r'C:\GITProject\pyschedule\nyk\File\Chemical_Unused.xlsx'
  wb.save(fullpath)

  conn.close()
  if P_ActiveQ1 > 0 or P_ActiveQ2 > 0:
    SendMailQN()
    print('Complete ExPortQnNoColor')


def SendMailQN():
  recipients = None
  P_REC = ""
  my_dsn = cx_Oracle.makedsn("172.16.6.76", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="demo", password="demo",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" Select Distinct 'A',SUPPORT_D2 From DFIT_SUPPORT Where Support_Grp = 'AUT' And Support_Id = 'MAIL' And SUPPORT_D1 = 'CHEMICAL' """)
  for A,SUPPORT_D2 in cursor:
    if P_REC is None or P_REC == "":
      P_REC = P_REC+SUPPORT_D2
    else:
      P_REC = P_REC+","+SUPPORT_D2

    print(P_REC)

  conn.close()
 

  msg = MIMEMultipart()
  message = "\r\r\n Thank you"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

  #recipients = "akawit.n@nanyangtextile.com"
    # bcc = 'piyawat.k@nanyangtextile.com,kaweewat.k@nanyangtextile.com,somchai.c@nanyangtextile.com'
    # recipients = 'kaweewat.k@nanyangtextile.com'
    # bcc = 'kaweewat.k@gmail.com'

  msg['From'] = "admin_monitoring.k@nanyangtextile.com"
  msg['To'] = P_REC
  msg['Subject'] = 'Auto Mail Warning Revise Formula.Finish Check Chemical Code Change Unused'
    
    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    # msg.attach(MIMEText(message, 'plain'))

  filename = r'Chemical_Unused.xlsx'
  attachment = open(r'C:\GITProject\pyschedule\nyk\File\Chemical_Unused.xlsx', 'rb')
    #attachment = open(r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx', 'rb')
  xlsx = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  xlsx.set_payload(attachment.read())

  encoders.encode_base64(xlsx)
  xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
  msg.attach(xlsx)

  rcpt = P_REC.split(',')

  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #erver.ehlo()  # Can be omitted
    #server.starttls(context=context)
  server.ehlo()  # Can be omitted
    #server.login('automail.nanyang@gmail.com', 'P@vilion1974')
  server.sendmail(msg['From'], rcpt, msg.as_string())
  server.quit()
  print("Send Mail")

    #try:
    #  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #  server.ehlo()  # Can be omitted
    #  server.starttls(context=context)
    #  server.ehlo()  # Can be omitted
    #  server.login('automail.nanyang@gmail.com', 'P@vilion1974')
    #  server.sendmail(msg['From'], rcpt, msg.as_string())
    #  server.quit()
    #  print("Send Mail")
    #except:
     #sys.exit("mail failed; %s" % "CUSTOM_ERROR")  # give an error message


############################################
threads = []

thread2 = CLS_ExPortQnNoColor()
#thread2 = CLS_SendMailExpire()
thread2.start()
threads.append(thread2)

for t in threads:
    t.join()
print("COMPLETE")