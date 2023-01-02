#-*-coding: utf-8 -*-
import tkinter as tk
from tkinter import messagebox
import cx_Oracle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import datetime
import threading
import smtplib
import ssl
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"


class CLS_ExPortQnNoColor(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      ExPortQnNoColor()

def ExPortQnNoColor():
  P_Active = 'N'
  FileNameXLSX = 'Dyestff_No_Orgx.xlsx'
  time_start = datetime.now()

  my_dsn = cx_Oracle.makedsn("172.16.6.76", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="demo", password="demo",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" select Items_Group,Items_Code,ITEMS_DESCRIP,Items_type,Items_Unit,Create_Date from V_Idyestff_No_ProdOrgx """)

  wb = openpyxl.Workbook()
  ws = wb['Sheet']

  #for Items_Group,Items_Code,ITEMS_DESCRIP,Items_type,Items_Unit,Create_Date in cursor:
  #  if Items_Code is not None:
  #    P_Active = "Y"

  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])

  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
      if row[c] is not None:
        P_Active = "Y"
    r = r + 1
  #fullpath = r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx'
  fullpath = r'C:\GITProject\pyschedule\nyk\File\Dyestff_No_Orgx.xlsx'
  wb.save(fullpath)

  conn.close()
  if P_Active == "Y":
    SendMailQN()
    print('Complete ExPortQnNoColor')


def SendMailQN():
  recipients = None
  P_REC = ""
  my_dsn = cx_Oracle.makedsn("172.16.6.76", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="demo", password="demo",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" Select Distinct 'A',SUPPORT_D2 From DFIT_SUPPORT Where Support_Grp = 'AUT' And Support_Id = 'MAIL' And SUPPORT_D1 = 'DYESTFF' """)
  for A,SUPPORT_D2 in cursor:
    if P_REC is None or P_REC == "":
      P_REC = P_REC+SUPPORT_D2
    else:
      P_REC = P_REC+","+SUPPORT_D2

  conn.close()
 

  msg = MIMEMultipart()
  message = "\r\r\n Thank you"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

  #recipients = "akawit.n@nanyangtextile.com"
    # bcc = 'piyawat.k@nanyangtextile.com,kaweewat.k@nanyangtextile.com,somchai.c@nanyangtextile.com'
    # recipients = 'kaweewat.k@nanyangtextile.com'
    # bcc = 'kaweewat.k@gmail.com'

  msg['From'] = "admin_monitoring.k@nanyangtextile.com"
  msg['To'] = P_REC
  msg['Subject'] = 'Auto Mail Warning Check Items Dyestff / Chemical No Prod Id Link Orgx'
    
    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    # msg.attach(MIMEText(message, 'plain'))

  filename = r'Dyestff_No_Orgx.xlsx'
  attachment = open(r'C:\GITProject\pyschedule\nyk\File\Dyestff_No_Orgx.xlsx', 'rb')
    #attachment = open(r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx', 'rb')
  xlsx = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  xlsx.set_payload(attachment.read())

  encoders.encode_base64(xlsx)
  xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
  msg.attach(xlsx)

  rcpt = P_REC.split(',')

  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #erver.ehlo()  # Can be omitted
    #server.starttls(context=context)
  server.ehlo()  # Can be omitted
    #server.login('automail.nanyang@gmail.com', 'P@vilion1974')
  server.sendmail(msg['From'], rcpt, msg.as_string())
  server.quit()
  print("Send Mail")

    #try:
    #  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #  server.ehlo()  # Can be omitted
    #  server.starttls(context=context)
    #  server.ehlo()  # Can be omitted
    #  server.login('automail.nanyang@gmail.com', 'P@vilion1974')
    #  server.sendmail(msg['From'], rcpt, msg.as_string())
    #  server.quit()
    #  print("Send Mail")
    #except:
     #sys.exit("mail failed; %s" % "CUSTOM_ERROR")  # give an error message


############################################
threads = []

thread2 = CLS_ExPortQnNoColor()
#thread2 = CLS_SendMailExpire()
thread2.start()
threads.append(thread2)

for t in threads:
    t.join()
print("COMPLETE")