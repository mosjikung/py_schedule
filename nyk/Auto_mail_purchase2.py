#-*-coding: utf-8 -*-
import tkinter as tk
from tkinter import messagebox
import cx_Oracle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import datetime
import threading
import smtplib
import ssl
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"
inDate = datetime.now()
P_Time1 = inDate.strftime("%d-%m-%Y")


class CLS_ExPortQnNoColor(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      ExPortQnNoColor()

def ExPortQnNoColor():
  P_ActiveQ1 = 0
  P_ActiveQ2 = 0
  FileNameXLSX = 'REQUEST_YARN_PRICE'+str(P_Time1)+'.xlsx'
  time_start = datetime.now()

  my_dsn = cx_Oracle.makedsn("172.16.6.74", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="sf5", password="omsf5",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" SELECT  REQUEST_DATE ,SUPPLIER_NAME,SUPPLIER_CODE,YARN_ITEM,YARN_COLOR,
                  PRICE_PER_KG,MIN_QTY,MAX_QTY,MOQ_KGS_SAMPLE,MOQ_KGS,PRODUCTION_LT,
                  TRANSIT_SEALT,TRANSIT_AIRLT	,WEIGHT_PERCON	,STANDARD_PACKAGE	,
                  WEIGHT_PERCORTON	,CARTON_TYPE
FROM FMIT_REQ_YARNPRICE
WHERE SEND_DATE IS NULL """)

  wb = openpyxl.Workbook()
  ws = wb['Sheet']

  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])
    P_ActiveQ1 = 1

  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
    r = r + 1
    P_ActiveQ2 = P_ActiveQ2 + 1
  #fullpath = r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx'
  fullpath = r'C:\GITProject\pyschedule\nyk\File\REQUEST_YARN_PRICE'+str(P_Time1)+'.xlsx'
  wb.save(fullpath)

  conn.close()
  if P_ActiveQ1 > 0 or P_ActiveQ2 > 0:
    SendMailQN()
    print('Complete ExPortQnNoColor')


def SendMailQN():
  recipients = None
  P_REC = ""
  my_dsn = cx_Oracle.makedsn("172.16.6.74", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="SF5", password="OMSF5",dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  cursor.execute(""" Select Distinct 'A',SUPPORT_D2 From DFIT_SUPPORT Where SUPPORT_GRP='SUB' AND SUPPORT_ID='AUTO' AND SUPPORT_D1='MAIL'""")
  for A,SUPPORT_D2 in cursor:
    if P_REC is None or P_REC == "":
      P_REC = P_REC+SUPPORT_D2
    else:
      P_REC = P_REC+","+SUPPORT_D2

  conn.close()
 

  msg = MIMEMultipart()
  message = "\r\r\n Thank you"
  body = "Dear,Sir \n \nThis is an Auto-email Request Yarn-Price,Please Full-fill information and reply to me. \nPlease see information in attach file.\n "
  #body = "This is an email with attachment sent from Python \nThank you \n test"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

  #recipients = "akawit.n@nanyangtextile.com"
    # bcc = 'piyawat.k@nanyangtextile.com,kaweewat.k@nanyangtextile.com,somchai.c@nanyangtextile.com'
    # recipients = 'kaweewat.k@nanyangtextile.com'
    # bcc = 'kaweewat.k@gmail.com'

  msg['From'] = "admin_monitoring.k@nanyangtextile.com"
  msg['To'] = P_REC
  msg['Subject'] = 'Auto-email Request Yarn-Price'
    
    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    # msg.attach(MIMEText(message, 'plain'))

  filename = r'REQUEST_YARN_PRICE'+str(P_Time1)+'.xlsx'
  attachment = open(r'C:\GITProject\pyschedule\nyk\File\REQUEST_YARN_PRICE'+str(P_Time1)+'.xlsx', 'rb')
    #attachment = open(r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx', 'rb')
    
  msg.attach(MIMEText(body, "plain"))
  xlsx = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  xlsx.set_payload(attachment.read())

  encoders.encode_base64(xlsx)
  xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
  msg.attach(xlsx)

  rcpt = P_REC.split(',')

  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #erver.ehlo()  # Can be omitted
    #server.starttls(context=context)
  server.ehlo()  # Can be omitted
    #server.login('automail.nanyang@gmail.com', 'P@vilion1974')
  server.sendmail(msg['From'], rcpt, msg.as_string())
  server.quit()
  print("Send Mail")

    #try:
    #  server = smtplib.SMTP('smtp.nanyangtextile.com: 25')
    #  server.ehlo()  # Can be omitted
    #  server.starttls(context=context)
    #  server.ehlo()  # Can be omitted
    #  server.login('automail.nanyang@gmail.com', 'P@vilion1974')
    #  server.sendmail(msg['From'], rcpt, msg.as_string())
    #  server.quit()
    #  print("Send Mail")
    #except:
     #sys.exit("mail failed; %s" % "CUSTOM_ERROR")  # give an error message


############################################
threads = []

thread2 = CLS_ExPortQnNoColor()
#thread2 = CLS_SendMailExpire()
thread2.start()
threads.append(thread2)

for t in threads:
    t.join()
print("COMPLETE")