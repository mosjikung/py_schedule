#-*-coding: utf-8 -*-
import tkinter as tk
from tkinter import messagebox
import cx_Oracle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import datetime
import threading
import smtplib
import ssl
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

oracle_client = "C:\instantclient_19_5"
os.environ["ORACLE_HOME"] = oracle_client
os.environ["PATH"] = oracle_client+os.pathsep+os.environ["PATH"]
os.environ["NLS_LANG"] = "AMERICAN_AMERICA.TH8TISASCII"

# def gen_cursor_to_file(FileNameXLSX, cursor):
#   wb = openpyxl.Workbook()
#   ws = wb['Sheet']
#   home = os.path.expanduser('~')

#   for c, col in enumerate(cursor.description):
#     ws.cell(row=1, column=c+1, value=col[0])

#   r = 2
#   for row in cursor:
#     for c, col in enumerate(cursor.description):
#       ws.cell(row=r, column=c+1, value=row[c])
#     r = r + 1
#   fullpath = home+'\\Downloads\\'+FileNameXLSX
#   wb.save(fullpath)


class CLS_ExPortQnNoColor(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      ExPortQnNoColor()

def ExPortQnNoColor():
  FileNameXLSX = 'QN-Data-NoColor.xlsx'
  time_start = datetime.now()
  my_dsn = cx_Oracle.makedsn("172.16.6.74", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="sf5", password="omsf5",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()

  cursor.execute(""" 
  SELECT QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
, PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
,ITEM_CODE
, LISTAGG(TOPDYED_COLOR, ',') WITHIN GROUP (ORDER BY TOPDYED_COLOR) TOPDYED_COLOR
, LISTAGG(ITEM_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_TYPE) ITEM_TYPE
, AVG(CONSUMPTION) CONSUMPTION
, LISTAGG(FABRIC_SHADE, ',') WITHIN GROUP (ORDER BY FABRIC_SHADE) FABRIC_SHADE
, LISTAGG(ITEM_PRICE_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_PRICE_TYPE) ITEM_PRICE_TYPE
, AVG(CONSUMPTION) ITEM_PRICE
, round(AVG(TARGET_PRICE),4) TARGET_PRICE
, round(AVG(ORIGINAL_PRICE),4) ORIGINAL_PRICE
, round(AVG(ITEM_CM_PRICE),4) ITEM_CM_PRICE
, round(AVG(STD_ITEM_PRICE),4) STD_ITEM_PRICE
, LISTAGG(LINE_REMARK, ',') WITHIN GROUP (ORDER BY LINE_REMARK) LINE_REMARK
, LISTAGG(SHADE_DESC, ',') WITHIN GROUP (ORDER BY SHADE_DESC) SHADE_DESC
, LISTAGG(SHADE_L, ',') WITHIN GROUP (ORDER BY SHADE_L) SHADE_L
, LISTAGG(SHADE_M, ',') WITHIN GROUP (ORDER BY SHADE_M) SHADE_M
, LISTAGG(SHADE_D, ',') WITHIN GROUP (ORDER BY SHADE_D) SHADE_D
, LISTAGG(SHADE_S, ',') WITHIN GROUP (ORDER BY SHADE_S) SHADE_S
, LISTAGG(USER_REMARK, ',') WITHIN GROUP (ORDER BY USER_REMARK) USER_REMARK
, LISTAGG(FIRST_BATCH, ',') WITHIN GROUP (ORDER BY FIRST_BATCH) FIRST_BATCH
, LISTAGG(MATCH_CONTRAST, ',') WITHIN GROUP (ORDER BY MATCH_CONTRAST) MATCH_CONTRAST
, LISTAGG(BODY_TRIM, ',') WITHIN GROUP (ORDER BY BODY_TRIM) BODY_TRIM
, LISTAGG(PILOT_RUN, ',') WITHIN GROUP (ORDER BY PILOT_RUN) PILOT_RUN
, LISTAGG(STYLE_REVIEW, ',') WITHIN GROUP (ORDER BY STYLE_REVIEW) STYLE_REVIEW
, SUM(ITEM_WEIGHT) ITEM_WEIGHT
, SUM(TOTAL_QTY) TOTAL_QTY
, LISTAGG(UOM, ',') WITHIN GROUP (ORDER BY UOM) UOM
, OPEN_QN('C'||QN_NO,ITEM_CODE) OPEN_DUMMY_SO
, MAX(RS_RESERVE) RS_RESERVE
, MAX(RS_ORDER) RS_ORDER
, MIN(EXPIRE_DATE) EXPIRE_DATE
, SUM(PENDING_QTY) PENDING_QTY
, SUM(DROP_QTY) DROP_QTY
, SUM(CLOSED_QTY) CLOSED_QTY
, SUM(RESERVE_QTY) RESERVE_QTY
,'ITEM_WEIGHT-OPEN_DUMMY_SO-DROP_QTY' FORMULA
, (SUM(ITEM_WEIGHT) - OPEN_QN('C'||QN_NO,ITEM_CODE) - SUM(DROP_QTY)) AS BALANCE_QTY
, DECODE(UNOPEN_DUMMY(QN_NO),'N','CONTRACT','') CONTRACT
, GET_PRE_QN_STATUS(QN_NO) SO_DUMMY_STATUS
, LISTAGG(CONFIRM_PLAN, ',') WITHIN GROUP (ORDER BY CONFIRM_PLAN) CONFIRM_PLAN
, SUM(COMFIRM_KGS) COMFIRM_KGS
, MIN(YEAR) YEAR, MIN(WEEK) WEEK, MIN(YEAR) || MIN(WEEK) AS WW
, LISTAGG(LINE_ID, ',') WITHIN GROUP (ORDER BY LINE_ID) LINE_ID
, LISTAGG(DROPPED_REASON, ',') WITHIN GROUP (ORDER BY DROPPED_REASON) DROPPED_REASON
, MIN(O_YARN_COUNT) O_YARN_COUNT, MIN(O_CONTENT) O_CONTENT
FROM (
SELECT QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
, PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
--, SUBSTR(ITEM_CODE,1,LENGTH(ITEM_CODE)-2) ITEM_CODE
, ITEM_CODE
, LISTAGG(TOPDYED_COLOR, ',') WITHIN GROUP (ORDER BY TOPDYED_COLOR) TOPDYED_COLOR
, LISTAGG(ITEM_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_TYPE) ITEM_TYPE
, AVG(CONSUMPTION) CONSUMPTION
, LISTAGG(FABRIC_SHADE, ',') WITHIN GROUP (ORDER BY FABRIC_SHADE) FABRIC_SHADE
, LISTAGG(ITEM_PRICE_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_PRICE_TYPE) ITEM_PRICE_TYPE
, AVG(CONSUMPTION) ITEM_PRICE
, AVG(TARGET_PRICE) TARGET_PRICE
, AVG(ORIGINAL_PRICE) ORIGINAL_PRICE
, AVG(ITEM_CM_PRICE) ITEM_CM_PRICE
, AVG(STD_ITEM_PRICE) STD_ITEM_PRICE
, LISTAGG(LINE_REMARK, ',') WITHIN GROUP (ORDER BY LINE_REMARK) LINE_REMARK
, LISTAGG(SHADE_DESC, ',') WITHIN GROUP (ORDER BY SHADE_DESC) SHADE_DESC
, LISTAGG(SHADE_L, ',') WITHIN GROUP (ORDER BY SHADE_L) SHADE_L
, LISTAGG(SHADE_M, ',') WITHIN GROUP (ORDER BY SHADE_M) SHADE_M
, LISTAGG(SHADE_D, ',') WITHIN GROUP (ORDER BY SHADE_D) SHADE_D
, LISTAGG(SHADE_S, ',') WITHIN GROUP (ORDER BY SHADE_S) SHADE_S
, LISTAGG(USER_REMARK, ',') WITHIN GROUP (ORDER BY USER_REMARK) USER_REMARK
, LISTAGG(FIRST_BATCH, ',') WITHIN GROUP (ORDER BY FIRST_BATCH) FIRST_BATCH
, LISTAGG(MATCH_CONTRAST, ',') WITHIN GROUP (ORDER BY MATCH_CONTRAST) MATCH_CONTRAST
, LISTAGG(BODY_TRIM, ',') WITHIN GROUP (ORDER BY BODY_TRIM) BODY_TRIM
, LISTAGG(PILOT_RUN, ',') WITHIN GROUP (ORDER BY PILOT_RUN) PILOT_RUN
, LISTAGG(STYLE_REVIEW, ',') WITHIN GROUP (ORDER BY STYLE_REVIEW) STYLE_REVIEW
, SUM(ITEM_WEIGHT) ITEM_WEIGHT
, SUM(TOTAL_QTY) TOTAL_QTY
, LISTAGG(UOM, ',') WITHIN GROUP (ORDER BY UOM) UOM
--OPEN_DUMMY_SO
--, GET_RS_QTY_RESERVED(QN_NO,ITEM_CODE) RS_RESERVE
--, GET_RS_QTY_ORDERED(QN_NO,ITEM_CODE) RS_ORDER
, MAX(RS_RESERVE) RS_RESERVE
, MAX(RS_ORDER) RS_ORDER
, TO_CHAR(EXPIRE_DATE,'YYYY-MM-DD') EXPIRE_DATE
, SUM(PENDING_QTY) PENDING_QTY
, SUM(DROP_QTY) DROP_QTY
, SUM(CLOSED_QTY) CLOSED_QTY
, SUM(RESERVE_QTY) RESERVE_QTY
,'ITEM_WEIGHT-OPEN_DUMMY_SO-DROP_QTY' FORMULA
--BALANCE_QTY
--, DECODE(UNOPEN_DUMMY(QN_NO),'N','CONTRACT','') CONTRACT
, GET_PRE_QN_STATUS(QN_NO) SO_DUMMY_STATUS
, LISTAGG(CONFIRM_PLAN, ',') WITHIN GROUP (ORDER BY CONFIRM_PLAN) CONFIRM_PLAN
, SUM(COMFIRM_KGS) COMFIRM_KGS
, YEAR, WEEK, YEAR || WEEK AS WW
, LINE_ID
,(SELECT MAX(DROPPED_REASON) FROM DFIT_BF_TARGET BF WHERE M.QN_NO = BF.ORDER_NO AND M.LINE_ID = BF.LINE_ID) DROPPED_REASON
, O_YARN_COUNT, O_CONTENT
FROM (
SELECT M.ORDER_NO QN_NO, TO_CHAR(M.ORDER_DATE,'YYYY-MM-DD') QN_DATE, M.ENTRY_TIME,M.CUSTOMER_ID, M.CUSTOMER_NAME, M.CUSTOMER_END, M.BF_TYPE, M.C_TYPE, M.SALES_ID, M.SALES_BY,  M.EXCHANGE_RATE, M.CURRENCY, M.PAYMENT_CONDITION
,M.QUOTATION_TYPE, M.PRICE_TYPE, M.COUNTRY, M.TEAM_NAME, M.CONVERT_QN, M.CONVERT_DATE, M.RQN_APPROVED AS QN_STATUS, M.RQN_DATE, M.RQN_BY, M.SEASON, M.STYLING, M.COMPLETE_QN, M.QN_TYPE, M.FREIGHT_COST
,D.LINE_ID, D.ITEM_CODE, D.TOPDYED_COLOR, D.ITEM_TYPE,  D.CONSUMPTION,  D.FABRIC_SHADE, D.PRICE_TYPE ITEM_PRICE_TYPE, D.ACTIVE_STATUS,  D.ITEM_PRICE ,D.TARGET_PRICE , D.ORIGINAL_PRICE, D.ITEM_CM_PRICE, D.STD_ITEM_PRICE
,D.LINE_REMARK , D.SHADE_DESC, D.SHADE_L, D.SHADE_M, D.SHADE_D, D.SHADE_S, D.USER_REMARK, D.FIRST_BATCH, D.MATCH_CONTRAST, D.BODY_TRIM , D.PILOT_RUN, D.STYLE_REVIEW  
,D.ITEM_WEIGHT, D.TOTAL_QTY, D.UOM
,(SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = M.ORDER_NO) EXPIRE_DATE
,GET_PRE_QN_LINE_PENDING(M.ORDER_NO,D.LINE_ID) PENDING_QTY
,GET_PRE_QN_LINE_DROP(M.ORDER_NO,D.LINE_ID) DROP_QTY
,GET_PRE_QN_LINE_CLOSED(M.ORDER_NO,D.LINE_ID) CLOSED_QTY
,GET_PRE_QN_LINE_RESERVED(M.ORDER_NO,D.LINE_ID) RESERVE_QTY
,CASE WHEN D.confirm_plan = 'CONFIRM' THEN d.item_weight ELSE 0 END COMFIRM_KGS
,D.CONFIRM_PLAN
,TO_CHAR(M.ORDER_DATE,'IYYY') YEAR
,TO_CHAR(M.ORDER_DATE,'IW') WEEK
, GET_RS_QTY_RESERVED(M.ORDER_NO, D.ITEM_CODE) RS_RESERVE
, GET_RS_QTY_ORDERED(M.ORDER_NO, D.ITEM_CODE) RS_ORDER
, I.O_YARN_COUNT, I.O_CONTENT
FROM DFIT_QN_REH M, DFIT_QN_RED D, SF5.FMIT_ITEM I 
WHERE M.ORDER_NO = D.ORDER_NO
AND M.ORDER_NO >= 618692
AND D.ITEM_CODE = I.ITEM_CODE(+) 
AND TO_CHAR(M.ORDER_DATE,'YYYY-MM-DD') >= '2019-01-01' 
AND NVL(M.RQN_APPROVED,'X') <> 'DELETE' 
--and M.ORDER_NO = 652006
--and d.ITEM_CODE =  'FD6GNTLH30/16A0'
) M
GROUP BY QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
,PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
,ITEM_CODE, YEAR, WEEK, TO_CHAR(EXPIRE_DATE,'YYYY-MM-DD'), EXPIRE_DATE, LINE_ID, O_YARN_COUNT, O_CONTENT
ORDER BY QN_DATE DESC, QN_NO DESC, ITEM_CODE )

GROUP BY QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
, PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
,ITEM_CODE
   """)



#   cursor.execute("""SELECT QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
# , PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
# , SUBSTR(ITEM_CODE,1,LENGTH(ITEM_CODE)-2) ITEM_CODE
# , LISTAGG(TOPDYED_COLOR, ',') WITHIN GROUP (ORDER BY TOPDYED_COLOR) TOPDYED_COLOR
# , LISTAGG(ITEM_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_TYPE) ITEM_TYPE
# , AVG(CONSUMPTION) CONSUMPTION
# , LISTAGG(FABRIC_SHADE, ',') WITHIN GROUP (ORDER BY FABRIC_SHADE) FABRIC_SHADE
# , LISTAGG(ITEM_PRICE_TYPE, ',') WITHIN GROUP (ORDER BY ITEM_PRICE_TYPE) ITEM_PRICE_TYPE
# , LISTAGG(ITEM_PRICE, ',') WITHIN GROUP (ORDER BY ITEM_PRICE) ITEM_PRICEcls
# , LISTAGG(TARGET_PRICE, ',') WITHIN GROUP (ORDER BY TARGET_PRICE) TARGET_PRICE
# , LISTAGG(ORIGINAL_PRICE, ',') WITHIN GROUP (ORDER BY ORIGINAL_PRICE) ORIGINAL_PRICE
# , LISTAGG(ITEM_CM_PRICE, ',') WITHIN GROUP (ORDER BY ITEM_CM_PRICE) ITEM_CM_PRICE
# , LISTAGG(STD_ITEM_PRICE, ',') WITHIN GROUP (ORDER BY STD_ITEM_PRICE) STD_ITEM_PRICE
# , LISTAGG(LINE_REMARK, ',') WITHIN GROUP (ORDER BY LINE_REMARK) LINE_REMARK
# , LISTAGG(SHADE_DESC, ',') WITHIN GROUP (ORDER BY SHADE_DESC) SHADE_DESC
# , LISTAGG(SHADE_L, ',') WITHIN GROUP (ORDER BY SHADE_L) SHADE_L
# , LISTAGG(SHADE_M, ',') WITHIN GROUP (ORDER BY SHADE_M) SHADE_M
# , LISTAGG(SHADE_D, ',') WITHIN GROUP (ORDER BY SHADE_D) SHADE_D
# , LISTAGG(SHADE_S, ',') WITHIN GROUP (ORDER BY SHADE_S) SHADE_S
# , LISTAGG(USER_REMARK, ',') WITHIN GROUP (ORDER BY USER_REMARK) USER_REMARK
# , LISTAGG(FIRST_BATCH, ',') WITHIN GROUP (ORDER BY FIRST_BATCH) FIRST_BATCH
# , LISTAGG(MATCH_CONTRAST, ',') WITHIN GROUP (ORDER BY MATCH_CONTRAST) MATCH_CONTRAST
# , LISTAGG(BODY_TRIM, ',') WITHIN GROUP (ORDER BY BODY_TRIM) BODY_TRIM
# , LISTAGG(PILOT_RUN, ',') WITHIN GROUP (ORDER BY PILOT_RUN) PILOT_RUN
# , LISTAGG(STYLE_REVIEW, ',') WITHIN GROUP (ORDER BY STYLE_REVIEW) STYLE_REVIEW
# , SUM(ITEM_WEIGHT) ITEM_WEIGHT
# , SUM(TOTAL_QTY) TOTAL_QTY
# , LISTAGG(UOM, ',') WITHIN GROUP (ORDER BY UOM) UOM
# , OPEN_QN('C'||QN_NO,ITEM_CODE) OPEN_DUMMY_SO
# --, GET_RS_QTY_RESERVED(QN_NO,ITEM_CODE) RS_RESERVE
# --, GET_RS_QTY_ORDERED(QN_NO,ITEM_CODE) RS_ORDER
# , MAX(RS_RESERVE) RS_RESERVE
# , MAX(RS_ORDER) RS_ORDER
# , TO_CHAR(EXPIRE_DATE,'YYYY-MM-DD') EXPIRE_DATE
# , SUM(PENDING_QTY) PENDING_QTY
# , SUM(DROP_QTY) DROP_QTY
# , SUM(CLOSED_QTY) CLOSED_QTY
# , SUM(RESERVE_QTY) RESERVE_QTY
# ,'ITEM_WEIGHT-OPEN_DUMMY_SO-DROP_QTY' FORMULA
# , (SUM(ITEM_WEIGHT) - OPEN_QN('C'||QN_NO,ITEM_CODE) - SUM(DROP_QTY)) AS BALANCE_QTY
# , DECODE(UNOPEN_DUMMY(QN_NO),'N','CONTRACT','') CONTRACT
# , GET_PRE_QN_STATUS(QN_NO) SO_DUMMY_STATUS
# , LISTAGG(CONFIRM_PLAN, ',') WITHIN GROUP (ORDER BY CONFIRM_PLAN) CONFIRM_PLAN
# , SUM(COMFIRM_KGS) COMFIRM_KGS
# , YEAR, WEEK, YEAR || WEEK AS WW
# , LINE_ID
# ,(SELECT MAX(DROPPED_REASON) FROM DFIT_BF_TARGET BF WHERE M.QN_NO = BF.ORDER_NO AND M.LINE_ID = BF.LINE_ID) DROPPED_REASON
# , O_YARN_COUNT, O_CONTENT
# FROM (
# SELECT M.ORDER_NO QN_NO, TO_CHAR(M.ORDER_DATE,'YYYY-MM-DD') QN_DATE, M.ENTRY_TIME,M.CUSTOMER_ID, M.CUSTOMER_NAME, M.CUSTOMER_END, M.BF_TYPE, M.C_TYPE, M.SALES_ID, M.SALES_BY,  M.EXCHANGE_RATE, M.CURRENCY, M.PAYMENT_CONDITION
# ,M.QUOTATION_TYPE, M.PRICE_TYPE, M.COUNTRY, M.TEAM_NAME, M.CONVERT_QN, M.CONVERT_DATE, M.RQN_APPROVED AS QN_STATUS, M.RQN_DATE, M.RQN_BY, M.SEASON, M.STYLING, M.COMPLETE_QN, M.QN_TYPE, M.FREIGHT_COST
# ,D.LINE_ID, D.ITEM_CODE, D.TOPDYED_COLOR, D.ITEM_TYPE,  D.CONSUMPTION,  D.FABRIC_SHADE, D.PRICE_TYPE ITEM_PRICE_TYPE, D.ACTIVE_STATUS,  D.ITEM_PRICE ,D.TARGET_PRICE , D.ORIGINAL_PRICE, D.ITEM_CM_PRICE, D.STD_ITEM_PRICE
# ,D.LINE_REMARK , D.SHADE_DESC, D.SHADE_L, D.SHADE_M, D.SHADE_D, D.SHADE_S, D.USER_REMARK, D.FIRST_BATCH, D.MATCH_CONTRAST, D.BODY_TRIM , D.PILOT_RUN, D.STYLE_REVIEW  
# ,D.ITEM_WEIGHT, D.TOTAL_QTY, D.UOM
# ,(SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = M.ORDER_NO) EXPIRE_DATE
# ,GET_PRE_QN_LINE_PENDING(M.ORDER_NO,D.LINE_ID) PENDING_QTY
# ,GET_PRE_QN_LINE_DROP(M.ORDER_NO,D.LINE_ID) DROP_QTY
# ,GET_PRE_QN_LINE_CLOSED(M.ORDER_NO,D.LINE_ID) CLOSED_QTY
# ,GET_PRE_QN_LINE_RESERVED(M.ORDER_NO,D.LINE_ID) RESERVE_QTY
# ,CASE WHEN D.confirm_plan = 'CONFIRM' THEN d.item_weight ELSE 0 END COMFIRM_KGS
# ,D.CONFIRM_PLAN
# ,TO_CHAR(M.ORDER_DATE,'IYYY') YEAR
# ,TO_CHAR(M.ORDER_DATE,'IW') WEEK
# , GET_RS_QTY_RESERVED(M.ORDER_NO, D.ITEM_CODE) RS_RESERVE
# , GET_RS_QTY_ORDERED(M.ORDER_NO, D.ITEM_CODE) RS_ORDER
# , I.O_YARN_COUNT, I.O_CONTENT
# FROM DFIT_QN_REH M, DFIT_QN_RED D, SF5.FMIT_ITEM I 
# WHERE M.ORDER_NO = D.ORDER_NO
# AND M.ORDER_NO >= 618692
# AND D.ITEM_CODE = I.ITEM_CODE(+) 
# AND TO_CHAR(M.ORDER_DATE,'YYYY-MM-DD') >= '2019-01-01' 
# AND NVL(M.RQN_APPROVED,'X') <> 'DELETE' 
# ) M
# GROUP BY QN_NO, QN_DATE, ENTRY_TIME, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END, BF_TYPE, C_TYPE, SALES_ID, SALES_BY, EXCHANGE_RATE, CURRENCY, PAYMENT_CONDITION, QUOTATION_TYPE
# ,PRICE_TYPE, COUNTRY, TEAM_NAME, CONVERT_QN, CONVERT_DATE, QN_STATUS, RQN_DATE, RQN_BY, SEASON, STYLING, COMPLETE_QN, QN_TYPE, FREIGHT_COST
# ,ITEM_CODE, YEAR, WEEK, TO_CHAR(EXPIRE_DATE,'YYYY-MM-DD'), EXPIRE_DATE, LINE_ID, O_YARN_COUNT, O_CONTENT
# ORDER BY QN_DATE DESC, QN_NO DESC, ITEM_CODE """)

  # , (SELECT TO_CHAR(EXPIRED_DATE,'YYYY-MM-DD') FROM DFIT_MORDER O WHERE O.ORDER_NO = QN_NO) EXPIRE_DATE

#   , CASE WHEN  CASE WHEN TRUNC(EXPIRE_DATE) <= TRUNC(SYSDATE)THEN (SUM(ITEM_WEIGHT) - OPEN_QN('C'||QN_NO,ITEM_CODE)) ELSE 0 END < 0 THEN 0
#   ELSE (SUM(ITEM_WEIGHT) - OPEN_QN('C'||QN_NO,ITEM_CODE)) END EXPIRE_KGS
# , CASE WHEN TRUNC(EXPIRE_DATE) <= TRUNC(SYSDATE)THEN 0 ELSE
#     CASE WHEN NVL(LISTAGG(CONFIRM_PLAN, ',') WITHIN GROUP (ORDER BY CONFIRM_PLAN),'X')='X' THEN SUM(ITEM_WEIGHT) ELSE 0 END
#  END NON_CONFIRM

  # gen_cursor_to_file(FileNameXLSX, cursor)

  wb = openpyxl.Workbook()
  ws = wb['Sheet']

  for c, col in enumerate(cursor.description):
    ws.cell(row=1, column=c+1, value=col[0])

  r = 2
  for row in cursor:
    for c, col in enumerate(cursor.description):
      ws.cell(row=r, column=c+1, value=row[c])
    r = r + 1
  fullpath = r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx'
  wb.save(fullpath)

  conn.close()
  SendMailQN()
  print('Complete ExPortQnNoColor')


def SendMailQN():
    msg = MIMEMultipart()
    # message = "Production send mail to \r\r\n" + mailto + "\r\r\n Thank you"
    message = "\r\r\n Thank you"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

    recipients = "patcharabhorn.c@nanyangtextile.com,natcha.t@nanyangtextile.com,kaweewat.k@nanyangtextile.com"
    # bcc = 'piyawat.k@nanyangtextile.com,kaweewat.k@nanyangtextile.com,somchai.c@nanyangtextile.com'
    # recipients = 'kaweewat.k@nanyangtextile.com'
    # bcc = 'kaweewat.k@gmail.com'

    msg['From'] = "admin_monitoring.k@nanyangtextile.com"
    msg['To'] = recipients
    msg['Subject'] = 'Auto Mail Excel QN'
    
    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    # msg.attach(MIMEText(message, 'plain'))

    filename = r'QN-Data-NoColor.xlsx'
    attachment = open(r'C:\GITProject\pyschedule\QN-Data-NoColor.xlsx', 'rb')
    xlsx = MIMEBase(
        'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    xlsx.set_payload(attachment.read())

    encoders.encode_base64(xlsx)
    xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(xlsx)

    rcpt = recipients.split(',')

    server = smtplib.SMTP('smtp.nanyangtextile.com: 25')

    server.sendmail(msg['From'], rcpt, msg.as_string())
    server.quit()
    print("Send Mail")


class CLS_SendMailExpire(threading.Thread):
    def __init__(self):
      threading.Thread.__init__(self)

    def run(self):
      SendMailExpire()


def SendMailExpire():
  FileNameXLSX = 'QN-Data-NoColor.xlsx'
  time_start = datetime.now()
  my_dsn = cx_Oracle.makedsn("172.16.6.74", port=1521, sid="NYTG")
  conn = cx_Oracle.connect(user="sf5", password="omsf5",
                           dsn=my_dsn, encoding="UTF-8", nencoding="UTF-8")
  cursor = conn.cursor()
  # sql = """SELECT ORDER_NO, ORDER_DATE ,(SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = REH.ORDER_NO) EXPIRE_DATE,
  # BF_TYPE, C_TYPE, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END,	SALES_ID,	SALES_BY,	CURRENCY,	PAYMENT_CONDITION,	QUOTATION_TYPE,	EXCHANGE_RATE,PRICE_TYPE,COUNTRY, TEAM_NAME,	CONVERT_QN,	CONVERT_DATE,	RQN_APPROVED,RQN_DATE,RQN_BY, AGEN_COMMISSION, SEASON, STYLING, QN_TYPE, QN_TYPE_SUB, FREIGHT_COST, 
  #                 NVL((SELECT MAX(E_MAIL) FROM DFIT_SALE_MAIL WHERE GROUP_TYPE = 'SEND OWNER' AND SALE_ID = REH.SALES_ID),'') MAIL_SALE
  #                 FROM DFIT_QN_REH REH
  #                 WHERE REH.ORDER_NO IN (
  #                 SELECT M.ORDER_NO
  #                 FROM (
  #                 SELECT I.ORDER_NO, I.ENTRY_DATE, I.ITEM_CODE, I.UOM, I.ITEM_WEIGHT, I.TOTAL_QTY, DP.ITEM_WEIGHT AS LINE_TOTAL, DP.PENDING_BF_QTY, DP.CLOSED_BF_QTY, DP.DROPPED_BF_QTY
  #                 ,(SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = I.ORDER_NO) EXPIRE_DATE
  #                 FROM DFIT_QN_RED I, DFIT_DORDER DP
  #                 WHERE I.ORDER_NO = DP.ORDER_NO
  #                 AND I.LINE_ID = DP.LINE_ID
  #                 AND I.UOM = 'YARD'
  #                 AND EXTRACT(YEAR FROM I.ENTRY_DATE)>= 2019
  #                 ORDER BY  I.ENTRY_DATE ) M
  #                 WHERE PENDING_BF_QTY > 0
  #                 AND TRUNC(EXPIRE_DATE) >= TRUNC(SYSDATE) + 1
  #                 AND TRUNC(EXPIRE_DATE) <= TRUNC(SYSDATE) + 3
  #                 )
  #                 ORDER BY SALES_ID, (SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = REH.ORDER_NO)"""

  sql = """ SELECT REH.ORDER_NO, TO_CHAR(ORDER_DATE,'YYYY-MM-DD') ORDER_DATE ,(SELECT TO_CHAR(EXPIRED_DATE,'YYYY-MM-DD') FROM DFIC_MORDER O WHERE O.ORDER_NO = REH.ORDER_NO) EXPIRE_DATE,
   C_TYPE, CUSTOMER_ID, CUSTOMER_NAME, CUSTOMER_END,	
  M.ITEM_CODE, M.ITEM_WEIGHT AS ITEM_WEIGHT_KG, M.TOTAL_QTY, M.UOM, M.PENDING_BF_QTY AS PENDING_BF_KG, M.CLOSED_BF_QTY AS CLOSED_BF_KG, M.DROPPED_BF_QTY AS DROPPED_BF_KG,
  BF_TYPE, SALES_ID,	SALES_BY,	CURRENCY,	PAYMENT_CONDITION,	QUOTATION_TYPE,	EXCHANGE_RATE,PRICE_TYPE,COUNTRY, TEAM_NAME,	CONVERT_QN,	CONVERT_DATE,	RQN_APPROVED,RQN_DATE,RQN_BY, AGEN_COMMISSION, SEASON, STYLING, QN_TYPE, QN_TYPE_SUB, FREIGHT_COST, 
                  NVL((SELECT MAX(E_MAIL) FROM DFIT_SALE_MAIL WHERE GROUP_TYPE = 'SEND OWNER' AND SALE_ID = REH.SALES_ID),'') MAIL_SALE
                  FROM (
                  SELECT I.ORDER_NO, I.ENTRY_DATE, I.ITEM_CODE, I.UOM, I.ITEM_WEIGHT, I.TOTAL_QTY, DP.ITEM_WEIGHT AS LINE_TOTAL, DP.PENDING_BF_QTY, DP.CLOSED_BF_QTY, DP.DROPPED_BF_QTY
                  ,(SELECT EXPIRED_DATE FROM DFIC_MORDER O WHERE O.ORDER_NO = I.ORDER_NO) EXPIRE_DATE
                  FROM DFIT_QN_RED I, DFIT_DORDER DP
                  WHERE I.ORDER_NO = DP.ORDER_NO
                  AND I.LINE_ID = DP.LINE_ID
                  AND I.UOM = 'YARD'
                  AND EXTRACT(YEAR FROM I.ENTRY_DATE)>= 2019
                  ORDER BY  I.ENTRY_DATE ) M, DFIT_QN_REH REH
                  WHERE PENDING_BF_QTY > 0
                  AND REH.ORDER_NO = M.ORDER_NO
                  AND TRUNC(EXPIRE_DATE) >= TRUNC(SYSDATE) + 1
                  AND TRUNC(EXPIRE_DATE) <= TRUNC(SYSDATE) + 3
                  ORDER BY SALES_ID, M.EXPIRE_DATE """

  dfData = pd.read_sql_query(sql, conn)

  dfSales = dfData.groupby(['SALES_ID','MAIL_SALE']).size().reset_index(name='counts')

  print(dfSales)

  num_format = "{:,}".format

  for index, row in dfSales.iterrows():
    print(row['MAIL_SALE'])
    dfSale = dfData[dfData['MAIL_SALE'] == row['MAIL_SALE']]
    dfSale.to_excel("{}.xlsx".format(row['SALES_ID']), sheet_name='Sheet_name_1',index=False)

    html = """<html>
              <body>
                <table style='border-collapse: collapse;font-family: Arial, Helvetica, sans-serif;'>
                <tr style='background-color: #4CAF50;border: 1px solid #ddd;padding: 8px;color: white;'>
                <th style='border: 1px solid #ddd;'>QN No.</th>
                <th style='border: 1px solid #ddd;'>QN Date</th>
                <th style='border: 1px solid #ddd;'>Expire Date</th>
                <th style='border: 1px solid #ddd;'>Customer</th>
                <th style='border: 1px solid #ddd;'>End Buyer</th>
                <th style='border: 1px solid #ddd;'>Item</th>
                <th style='border: 1px solid #ddd;'>Weight KGs.</th>
                <th style='border: 1px solid #ddd;'>Quantity</th>
                <th style='border: 1px solid #ddd;'>UOM</th>
                <th style='border: 1px solid #ddd;'>Pending</th>
                <th style='border: 1px solid #ddd;'>Closed</th>
                <th style='border: 1px solid #ddd;'>Drop</th>
                </tr>
          """
    for index, row2 in dfSale.iterrows():
      html = html + """<tr>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;text-align: right;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;text-align: right;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;text-align: right;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;text-align: right;'>{}</td>
                    <td style='border: 1px solid #ddd;padding: 8px;text-align: right;'>{}</td>
                    </tr>""".format(row2['ORDER_NO'], row2['ORDER_DATE'], row2['EXPIRE_DATE'], row2['CUSTOMER_NAME'], row2['CUSTOMER_END'], row2['ITEM_CODE'], num_format(row2['ITEM_WEIGHT_KG']), num_format(row2['TOTAL_QTY']), row2['UOM'], num_format(row2['PENDING_BF_KG']), num_format(row2['CLOSED_BF_KG']), num_format(row2['DROPPED_BF_KG']))

    html = html + """ </table></body></html>"""

    SendMailAlertExpire(row['SALES_ID'], row['MAIL_SALE'], html)
    # print(dfSale.shape)


def SendMailAlertExpire(saleid, salemail, html):
    msg = MIMEMultipart()
    # message = "Production send mail to \r\r\n" + mailto + "\r\r\n Thank you"
    # _path = "C:\\GITProject\\pyschedule\\QN-Data-NoColor.xlsx"

    to = salemail
    to = 'kaweewat.k@nanyangtextile.com'
    cc = 'kaweewat.k@nanyangtextile.com'

    rcpt = cc.split(",") + [to]

    msg['From'] = "automail.nanyang@gmail.com"
    msg['To'] = to
    msg['Cc'] = cc
    msg['Subject'] = 'Auto Mail Expire QN of Sale {}'.format(saleid)
    
    # att = MIMEApplication(open("QN-Data-NoColor.xlsx", "rb").read())
    # msg.attach(att)
    part2 = MIMEText(html, "html")
    msg.attach(part2)
    # msg.attach(MIMEText("\r\r\n Thank you", 'plain'))

    filename = "{}.xlsx".format(saleid)
    attachment = open(filename, 'rb')
    xlsx = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    xlsx.set_payload(attachment.read())

    encoders.encode_base64(xlsx)
    xlsx.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(xlsx)

    # rcpt = recipients.split(',')
    context = ssl.create_default_context()

    try:
      # server = smtplib.SMTP('smtp.sinetFTTx.com: 25')
      # server.set_debuglevel(False)
      # server.login('admin_recruit@simat.co.th', 'Admin@1234')

      server = smtplib.SMTP('smtp.gmail.com', 587)
      server.ehlo()  # Can be omitted
      server.starttls(context=context)
      server.ehlo()  # Can be omitted
      server.login('automail.nanyang@gmail.com', 'P@vilion1974')
      server.sendmail(msg['From'], rcpt, msg.as_string())
      server.quit()
      print("Send Mail")
    except:
      sys.exit("mail failed; %s" % "CUSTOM_ERROR")  # give an error message

    # server = smtplib.SMTP('smtp.nanyangtextile.com: 25')

    # server.sendmail(msg['From'], rcpt, msg.as_string())
    # server.quit()
    print("Send Mail")

############################################
threads = []

thread1 = CLS_ExPortQnNoColor()
thread1.start()
threads.append(thread1)

# thread2 = CLS_SendMailExpire()
# thread2.start()
# threads.append(thread2)


for t in threads:
    t.join()
print("COMPLETE")



